from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from eth_utils import is_address, to_checksum_address
from fastapi import UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from .config import SecretBox
from .models import Chain, Job, JobItem, WalletImport, WalletImportMember
from .repositories import WalletRepository, log_event


class ImportIssue(BaseModel):
    kind: str
    source: str
    row: int
    value: str
    detail: str


class PreviewEntry(BaseModel):
    address: str
    checksum_address: str
    source: str
    row: int
    source_index: str | None = None


class ImportPreview(BaseModel):
    token: str
    valid_count: int
    duplicate_count: int
    invalid_count: int
    source_count: int
    entries: list[PreviewEntry]
    issues: list[ImportIssue]


class ConfirmImport(BaseModel):
    token: str
    name: str = Field(min_length=1, max_length=255)
    chain: str = "ethereum"


def normalize_address(value: object) -> tuple[str, str]:
    raw = str(value).strip()
    if not is_address(raw):
        raise ValueError("Некорректный адрес Ethereum")
    checksum = to_checksum_address(raw)
    return checksum.lower(), checksum


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1251"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Кодировка текста не поддерживается")


def _rows_from_csv(name: str, data: bytes) -> list[dict]:
    text = _decode_text(data)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [
        str(item or "").strip().lower() for item in (reader.fieldnames or [])
    ]
    index_headers = [item for item in headers if item != "wallet_address"]
    if (
        len(headers) != 2
        or "wallet_address" not in headers
        or len(index_headers) != 1
        or index_headers[0] not in {"", "index"}
    ):
        raise ValueError(
            f"{name}: нужны только столбцы index и 'wallet_address'"
        )
    index_header = index_headers[0]
    rows = []
    for row_number, row in enumerate(reader, start=2):
        normalized = {
            str(key or "").strip().lower(): value for key, value in row.items()
        }
        rows.append(
            {
                "value": normalized.get("wallet_address", ""),
                "index": normalized.get(index_header),
                "row": row_number,
            }
        )
    return rows


def _rows_from_xlsx(name: str, data: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator, ())
    normalized_headers = [
        str(value or "").strip().lower() for value in headers
    ]
    index_headers = [
        item for item in normalized_headers if item != "wallet_address"
    ]
    if (
        len(normalized_headers) != 2
        or "wallet_address" not in normalized_headers
        or len(index_headers) != 1
        or index_headers[0] not in {"", "index"}
    ):
        workbook.close()
        raise ValueError(
            f"{name}: нужны только столбцы index и 'wallet_address'"
        )
    positions = {
        str(value or "").strip().lower(): index
        for index, value in enumerate(headers)
    }
    index_header = index_headers[0]
    rows = []
    for row_number, values in enumerate(iterator, start=2):
        address_pos = positions["wallet_address"]
        index_pos = positions[index_header]
        rows.append(
            {
                "value": values[address_pos] if address_pos < len(values) else "",
                "index": (
                    str(values[index_pos])
                    if index_pos is not None
                    and index_pos < len(values)
                    and values[index_pos] is not None
                    else None
                ),
                "row": row_number,
            }
        )
    workbook.close()
    return rows


def _rows_from_text(data: bytes) -> list[dict]:
    return [
        {"value": line.strip(), "index": None, "row": row_number}
        for row_number, line in enumerate(_decode_text(data).splitlines(), start=1)
        if line.strip()
    ]


async def preview_import(
    files: list[UploadFile],
    manual_text: str,
    secret_box: SecretBox,
) -> ImportPreview:
    sources: list[tuple[str, list[dict]]] = []
    for upload in files:
        name = Path(upload.filename or "upload.txt").name
        data = await upload.read()
        suffix = Path(name).suffix.lower()
        if suffix == ".csv":
            rows = _rows_from_csv(name, data)
        elif suffix in {".xlsx", ".xlsm"}:
            rows = _rows_from_xlsx(name, data)
        elif suffix == ".txt":
            rows = _rows_from_text(data)
        else:
            raise ValueError(f"{name}: неподдерживаемый тип файла")
        sources.append((name, rows))
    if manual_text.strip():
        sources.append(("manual", _rows_from_text(manual_text.encode("utf-8"))))
    if not sources:
        raise ValueError("Добавьте хотя бы один файл или адрес вручную")

    entries: list[PreviewEntry] = []
    issues: list[ImportIssue] = []
    seen: dict[str, PreviewEntry] = {}
    for source, rows in sources:
        for row in rows:
            raw = str(row["value"] or "").strip()
            try:
                address, checksum = normalize_address(raw)
            except ValueError as exc:
                issues.append(
                    ImportIssue(
                        kind="invalid",
                        source=source,
                        row=row["row"],
                        value=raw,
                        detail=str(exc),
                    )
                )
                continue
            if address in seen:
                first = seen[address]
                issues.append(
                    ImportIssue(
                        kind="duplicate",
                        source=source,
                        row=row["row"],
                        value=raw,
                        detail=f"Первое вхождение: {first.source}:{first.row}",
                    )
                )
                continue
            entry = PreviewEntry(
                address=address,
                checksum_address=checksum,
                source=source,
                row=row["row"],
                source_index=row["index"],
            )
            seen[address] = entry
            entries.append(entry)

    payload = json.dumps(
        {"version": 1, "entries": [entry.model_dump() for entry in entries]},
        separators=(",", ":"),
    )
    token = secret_box.seal_json(payload)
    return ImportPreview(
        token=token,
        valid_count=len(entries),
        duplicate_count=sum(issue.kind == "duplicate" for issue in issues),
        invalid_count=sum(issue.kind == "invalid" for issue in issues),
        source_count=len(sources),
        entries=entries,
        issues=issues,
    )


def confirm_import(
    session: Session,
    request: ConfirmImport,
    secret_box: SecretBox,
) -> tuple[WalletImport, Job]:
    try:
        payload = json.loads(secret_box.open_json(request.token))
    except (ValueError, json.JSONDecodeError) as exc:
        raise ValueError("Некорректный или нечитаемый токен предпросмотра") from exc
    if payload.get("version") != 1 or not isinstance(payload.get("entries"), list):
        raise ValueError("Неподдерживаемый токен предпросмотра")
    if not payload["entries"]:
        raise ValueError("В импорте нет корректных адресов кошельков")

    chain = session.scalar(select(Chain).where(Chain.slug == request.chain))
    if chain is None or not chain.enabled:
        raise ValueError("Сеть неизвестна или отключена")

    batch = WalletImport(
        chain_id=chain.id,
        name=request.name,
        source_summary={
            "sources": sorted({entry["source"] for entry in payload["entries"]})
        },
        wallet_count=len(payload["entries"]),
    )
    session.add(batch)
    session.flush()
    wallets = WalletRepository(session)
    members = []
    for entry in payload["entries"]:
        normalized, checksum = normalize_address(entry["address"])
        wallet = wallets.get_or_create(chain.id, normalized, checksum)
        member = WalletImportMember(
            wallet_import_id=batch.id,
            wallet_id=wallet.id,
            source_name=str(entry["source"])[:255],
            source_row=int(entry["row"]),
            source_index=(
                str(entry["source_index"])[:255]
                if entry.get("source_index") is not None
                else None
            ),
        )
        session.add(member)
        members.append(member)

    job = Job(
        kind="collection",
        state="queued",
        wallet_import_id=batch.id,
        progress_total=len(members),
        parameters={"chain": chain.slug},
    )
    session.add(job)
    session.flush()
    for member in members:
        session.add(JobItem(job_id=job.id, wallet_id=member.wallet_id))
    log_event(
        session,
        "info",
        "import.confirmed",
        f"Created import with {len(members)} wallets",
        job_id=job.id,
        context={"wallet_import_id": batch.id},
    )
    session.flush()
    return batch, job
